from pyzabbix import ZabbixAPI
from openpyxl import *
from sys import argv


class Parser:
    def __init__(self):
        _, self.templateids = argv
        self.z = ZabbixAPI('https://zabbix.360on.ru', user='lldp', password='Roov2AiPh9ee')
        self.items = self.z.item.get(templateids=int(self.templateids))
        self.trigger = ""
        self.excel_cells = ['itemid', 'name', 'key_', 'history', 'trends', 'status', 'units', 'description']
        self.excel_cells_triggers = ['triggerid', 'expression', 'description','url', 'status', 'comments']
        self.macro = self.z.usermacro.get(hostids=self.templateids)

        self.prototype_cells = ['itemid', 'name', 'key_', 'history', 'trends', 'status', 'units', 'description']
        self.prototype_cells_triggers = ['triggerid', 'expression', 'description','url', 'status', 'comments']
        self.discoveryrule_cells = ['itemid', 'key_', 'name', 'status', 'description']

        self.itemprototype = self.z.itemprototype.get(templateids = int(self.templateids))
        self.triggerprototype = self.z.triggerprototype.get(templateids = int(self.templateids))

        self.discoveryrule = self.z.discoveryrule.get(templateids = int(self.templateids))
        # print(self.triggerprototype)
        # print(self.avg)

    def excel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'TemplateID ' + str(self.templateids)

        itemids = {}

        # получаем itemids и строчку в которой этот itemid находится
        for row in range(len(self.items)):
            column = 1
            for keys in self.items[row].keys():
                if keys == 'itemid':
                    itemids[str(self.items[row][keys])] = row+2

        itemids_with_trigger = []
        # print(itemids_with_trigger)
        keys_of_template = list(self.items[0].keys())

        for input_titles in self.excel_cells:
            if not input_titles in keys_of_template:
                print(input_titles + " is no in templates")
                self.excel_cells.remove(input_titles)


        column = 1
        column_trigger = 1
        for item in self.excel_cells:
            sheet.cell(row=1, column=column).value = item
            column += 1
        column_trigger = column
        for item in self.excel_cells_triggers:
            sheet.cell(row=1, column=column_trigger).value = item
            column_trigger += 1

        row = 1
        for items in self.items:
            column = 1
            row += 1
            for item in items:
                if item in self.excel_cells:
                    sheet.cell(row = row, column=column).value = items[item]
                    column += 1
                    if item == 'itemid':
                        triggers = self.z.trigger.get(templateids=self.templateids, itemids=items[item])
                        if triggers:
                            if len(triggers) > 1:
                                for x in range(len(triggers)):
                                    # print(items[item])
                                    row += 1
                                    row_for_triggers = row
                                    column_for_trigger = column_trigger -(len(self.excel_cells_triggers))
                                    for keys in triggers[x]:
                                        if keys in self.excel_cells_triggers:
                                                # print(keys + " is " + str(column_for_trigger))
                                                sheet.cell(row = row_for_triggers-1, column = column_for_trigger).value = triggers[x][keys]
                                                column_for_trigger += 1
                                row -= 1


                            else:
                                for trigger in triggers:
                                    # print(trigger)
                                    column_for_trigger = column_trigger-(len(self.excel_cells_triggers))
                                    for keys in trigger:
                                        if keys in self.excel_cells_triggers:
                                                # print(keys + " is " + str(column))
                                                sheet.cell(row = row, column = column_for_trigger).value = trigger[keys]
                                                column_for_trigger += 1

        sheet1 = wb.create_sheet('Usermacro')
        row = 1
        column = 1
        if self.macro:
            for macro in self.macro[0].keys():
                if macro == 'hostid':
                    sheet1.cell(row=row, column=column).value = macro
                    column += 1
                elif macro == 'macro':
                    sheet1.cell(row=row, column=column).value = macro
                    column += 1
                elif macro == 'value':
                    sheet1.cell(row=row, column=column).value = macro
                    column += 1
                elif macro == 'description':
                    sheet1.cell(row=row, column=column).value = macro
                    column += 1
                else:
                    continue

        row = 1
        if self.macro:
            for macros in self.macro:
                row += 1
                column = 1
                for macro in macros:
                    if macro == "hostid":
                        sheet1.cell(row=row, column=column).value = macros[macro]
                        column += 1
                    elif macro == 'macro':
                        sheet1.cell(row=row, column=column).value = macros[macro]
                        column += 1
                    elif macro == 'value':
                        sheet1.cell(row=row, column=column).value = macros[macro]
                        column += 1
                    elif macro == 'description':
                        sheet1.cell(row=row, column=column).value = macros[macro]
                        column += 1
                    else:
                        continue


        if self.itemprototype:
            sheet2 = wb.create_sheet('Prototypes')

            column = 1
            column_trigger = 1
            for item in self.prototype_cells:
                sheet2.cell(row=1, column=column).value = item
                column += 1
            column_trigger = column
            for item in self.prototype_cells_triggers:
                sheet2.cell(row=1, column=column_trigger).value = item
                column_trigger += 1

            row = 2
            itemids = {} # получаем itemid и строка в которой он находится
            # print(len(self.itemprototype))
            for items in self.itemprototype:
                for item in items:
                    if item in self.prototype_cells:
                        if item == 'itemid':
                            itemids[str(items[item])] = row
                            row += 1
            # print(itemids)

            itemids_with_trigger = []

            for i in itemids.keys():
                triggerprototype = self.z.triggerprototype.get(templateids = int(self.templateids), itemids = int(i))
                if triggerprototype:
                    itemids_with_trigger.append(i)
            row = 1

            for items in self.itemprototype:
                column = 1
                row += 1
                for item in items:
                    if item in self.prototype_cells:
                        sheet2.cell(row = row, column=column).value = items[item]
                        column += 1
                        if item == 'itemid':
                            triggers = self.z.triggerprototype.get(templateids=self.templateids, itemids=items[item])
                            if triggers:
                                if len(triggers) > 1:
                                    for x in range(len(triggers)):
                                        # print(items[item])
                                        row += 1
                                        row_for_triggers = row
                                        column_for_trigger = column_trigger -(len(self.prototype_cells_triggers))
                                        for keys in triggers[x]:
                                            if keys in self.prototype_cells_triggers:
                                                    # print(keys + " is " + str(column_for_trigger))
                                                    sheet2.cell(row = row_for_triggers-1, column = column_for_trigger).value = triggers[x][keys]
                                                    column_for_trigger += 1
                                    row -= 1


                                else:
                                    for trigger in triggers:
                                        # print(trigger)
                                        column_for_trigger = column_trigger-(len(self.prototype_cells_triggers))
                                        for keys in trigger:
                                            if keys in self.prototype_cells:
                                                    # print(keys + " is " + str(column))
                                                    sheet2.cell(row = row, column = column_for_trigger).value = trigger[keys]
                                                    column_for_trigger += 1

        if self.discoveryrule:
            column = 1
            sheet3 = wb.create_sheet('Discoveryrule')
            for item in self.discoveryrule_cells:
                sheet3.cell(row=1, column=column).value = item
                column += 1
            row = 1
            for items in self.discoveryrule:
                column = 1
                row += 1
                for item in items:
                    if item in self.discoveryrule_cells:
                        sheet3.cell(row = row, column=column).value = items[item]
                        column += 1

        wb.save('TemplateID ' + str(self.templateids) + '.xlsx')


if __name__ == "__main__":
    parser = Parser()
    parser.excel()
